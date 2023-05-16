import pandas as pd
from openpyxl import Workbook
import sys
import json


def getListsFromFile():
    # загружаем файл, переданный программе в ее аргументах
    file = load_file()

    # просим пользователя выбрать лист для работы программы
    lists = select_list(file)
    # на этом моменте получили целый лист, готовый к парсингу (selected_list)

    return lists


def saveTimetableToFile(timetables):
    wb = Workbook()

    for timetable in timetables:
        ws = wb.create_sheet(timetable)

        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 50

        timetable = timetables[timetable]
        ws.cell(row=1, column=1).value = 'Чётная неделя'
        for idx, day in enumerate(timetable['even']):
            for i, lesson in enumerate(day):
                if lesson == '':
                    continue
                else:
                    ws.cell(row=i + 2,
                            column=idx + 1).value = f'{i + 1}. {lesson["lesson"]} / {lesson["teacher"]} / {lesson["cabinet"]}'
        ws.cell(row=14, column=1).value = 'Чётная неделя'
        for idx, day in enumerate(timetable['odd']):
            for i, lesson in enumerate(day):
                if lesson == '':
                    continue
                else:
                    ws.cell(row=i + 15,
                            column=idx + 1).value = f'{i + 1}. {lesson["lesson"]} / {lesson["teacher"]} / {lesson["cabinet"]}'

    del wb['Sheet']
    wb.save("result.xlsx")
    wb.close()


def saveTimetableToJson(timetables, failed_lessons):
    result = {}
    result['timetables'] = timetables
    result['failed_lessons'] = failed_lessons
    return json.dumps(result)


def select_list(file):
    # ввод листа от пользователя
    # selected_list = 0
    # if (len(xl.sheet_names) > 1):
    #     print("Обнаружено несколько листов. Выберите необходимый лист:")
    #     # Печатаем название листов в данном файле
    #     for idx, list_name in enumerate(xl.sheet_names):
    #         print(f"{idx+1}. {list_name}")
    #     selected_list = input("Введите число: ")

    # лист с кабинетами
    cabinets_list = 0  # первый по счету лист
    cabinets_list = parse_list(file, cabinets_list)

    # лист с парами
    lessons_list = 1  # второй по счету лист
    lessons_list = parse_list(file, lessons_list)

    # лист с заказными парами
    ordered_lessons_list = 2  # третий по счету лист
    ordered_lessons_list = parse_list(file, ordered_lessons_list)

    return [cabinets_list, lessons_list, ordered_lessons_list]


def parse_list(file, selected_list):
    try:
        selected_list = int(selected_list)
        # Загрузить лист по его имени
        return file.parse(file.sheet_names[selected_list])
    except:
        print("Лист некорректен!")
        exit()


def load_file():
    # проверяем наличие переданного файла
    if (len(sys.argv) < 2):
        print("Ошибка: запуск без аргументов!")
        exit()
    else:
        in_file = sys.argv[1]

    # Загружаем spreadsheet в объект pandas
    file = pd.ExcelFile(in_file)
    return file